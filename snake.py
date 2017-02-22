from Tkinter import *
import tkMessageBox
import random
import threading
import time

import sys
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter


game_height = 15
game_width = 15

"""
 --------> x
 |
 |
 |
\|/
 y
"""


class Background():
    def __init__(self):
        self.root = Tk()

    def draw(self):
        self.cv = Canvas(self.root, bg='white', height=20 * game_height, width=20 * game_width)

        self.cv.pack()


class Snake():
    def __init__(self):
        # initial position(20,20),grid_size:20, direction: 0 and 1 are horizontal, 2 and 3 are vertical
        self.x = 20
        self.y = 20
        self.direction = 3
        self.number = 4

        # { {x0,y0}, {x1,y1}, ...,{xn,tn} }
        self.snake_coord = []
        self.block = []
        for i in range(self.number):
            coord = [self.x + 20, self.y]
            self.x = self.x + 20
            self.snake_coord.append(coord)
            rt = background.cv.create_rectangle(self.x, self.y, self.x + 20, self.y + 20, fill='blue')
            self.block.append(rt)
            background.cv.pack()


    def eat(self):
        coord = self.snake_coord[self.number - 1]
        self.snake_coord.append(coord)
        rt = background.cv.create_rectangle(coord[0], coord[1], coord[0] + 20, coord[1] + 20, fill='blue')
        self.block.append(rt)
        background.cv.pack()
        self.number = self.number + 1

    def move(self):
        # draw other blocks
        for i in range(1, self.number):
            self.snake_coord[self.number - i] = self.snake_coord[self.number - i - 1]
            x = self.snake_coord[self.number - i][0]
            y = self.snake_coord[self.number - i][1]
            #            if self.number - i == 1:
            #                print "test: ", self.snake_coord[self.number - i ]

            #            print self.number - i, self.snake_coord[self.number - i]
            #            print self.number - i - 1, self.snake_coord[self.number - i - 1]
            background.cv.coords(self.block[self.number - i], (x, y, x + 20, y + 20))

        x = self.snake_coord[0][0]
        y = self.snake_coord[0][1]
        if self.direction == 0:
            x = x + 20
        elif self.direction == 1:
            x = x - 20
        elif self.direction == 2:
            y = y - 20
        elif self.direction == 3:
            y = y + 20
        temp = [x, y]
        self.snake_coord[0] = temp
        # draw the first block
        background.cv.coords(self.block[0], (x, y, x + 20, y + 20))
        self.checkConflict()

    def checkConflict(self):
        # check wall, only with the head
        x = self.snake_coord[0][0]
        y = self.snake_coord[0][1]
        if x < 0 or x == 20 * game_width or y < 0 or y == 20 * game_height:
            print "hit the wall"
            tkMessageBox.showerror("Game Over", "You hit the wall")

        # check the body
        for i in range(1, self.number):
            if x == self.snake_coord[i][0] and y == self.snake_coord[i][1]:
                print "hit the body", i, x, " ", self.snake_coord[i][0], " ", y, self.snake_coord[i][1]
                tkMessageBox.showerror("Game Over", "You hit the body")

    def autoMove(self, goalCoord):
        startCoord = self.snake_coord[0]
        obstacleArray = []
        for i in range(1, len(self.snake_coord)):
            obstacleArray.append(self.snake_coord[i])
        pathplan = Pathplan(startCoord, goalCoord, obstacleArray)
        optimalPath = pathplan.findPath()
        i = len(optimalPath) - 1
        while i > 0 or i == 0:
            #            background.cv.create_rectangle(optimalPath[i][0], optimalPath[i][1], optimalPath[i][0]+20, optimalPath[i][1]+20, fill = 'red')
            time.sleep(0.5)
            #            print "path ", len(optimalPath) - i - 1, "  :", optimalPath[i]
            if optimalPath[i][0] > self.snake_coord[0][0]:
                self.direction = 0
            elif optimalPath[i][0] < self.snake_coord[0][0]:
                self.direction = 1
            elif optimalPath[i][1] > self.snake_coord[0][1]:
                self.direction = 3
            elif optimalPath[i][1] < self.snake_coord[0][1]:
                self.direction = 2
            self.move()
            i = i - 1


class Food():
    def __init__(self):
        x = 20 * random.randint(1, 10)
        y = 20 * random.randint(1, 10)
        self.coord = [x, y]
        self.rt = background.cv.create_rectangle(x, y, x + 20, y + 20, fill='green')
        background.cv.pack()

    def eaten(self, snake_body):
        # random
        inBody = 1
        number = len(snake_body)
        while(inBody == 1):
            inBody = 0
            x = 20 * random.randint(1, 14)
            y = 20 * random.randint(1, 14)
            #check if food colide with snake
            for i in range(number):
                if x == snake_body[i][0] and y == snake_body[i][1]:
                    inBody = 1
                    break
        self.coord = [x, y]
        background.cv.coords(self.rt, (x, y, x + 20, y + 20))


class SnakeGame():
    def __init__(self):

        self.snake = Snake()
        self.food = Food()
        self.entry = Entry(background.root)

        # create a datasheet to save data
        self.wb = Workbook()
        self.ws = self.wb.worksheets[0]
        self.ws.title = "trainning data 1"
        self.data_number = 1
        #Use button to save file
        Button(background.root, text = "save training data", command = self.saveFile).pack()

        threads = []
        t1 = threading.Thread(target=self.move)
        t1.start()
        threads.append(t1)

        t2 = threading.Thread(target=self.listenKeyboard)
 #       t2 = threading.Thread(target = self.autoMove, args = (self.food.coord, ) )
        t2.start()
        threads.append(t2)
        background.root.mainloop()

    def saveFile(self):
        self.wb.save("Training_data.xlsx")

    def changedirection(self, event):
        # print('You have entered: ', event.char)
        # use "w,a,s,d" to control direction
        time.sleep(0.3)
        self.writeStatus()
        if event.char == 'w':
            if self.snake.direction == 0 or self.snake.direction == 1:
                self.snake.direction = 2
                self.ws.cell(row=self.data_number, column=626).value = 1
        if event.char == 's':
            if self.snake.direction == 0 or self.snake.direction == 1:
                self.snake.direction = 3
                self.ws.cell(row=self.data_number, column=626).value = 2
        if event.char == 'd':
            if self.snake.direction == 2 or self.snake.direction == 3:
                self.snake.direction = 0
                self.ws.cell(row=self.data_number, column=626).value = 3
        if event.char == 'a':
            if self.snake.direction == 2 or self.snake.direction == 3:
                self.snake.direction = 1
                self.ws.cell(row=self.data_number, column=626).value = 4
        self.data_number = self.data_number + 1

    def listenKeyboard(self):

        self.entry.bind('<Key>', self.changedirection)
        self.entry.pack()

    def move(self):
        while (1):
            time.sleep(0.3)
            self.writeStatus()
            self.ws.cell(row=self.data_number, column = 626).value = 0
            self.data_number = self.data_number + 1
            self.snake.move()
            if self.snake.snake_coord[0] == self.food.coord:
                print "eat"
                self.snake.eat()
                self.food.eaten(self.snake.snake_coord)
                goalCoord = self.food.coord
                #time.sleep(0.2)

    def autoMove(self, goalCoord):
        self.snake.autoMove(goalCoord)

    #        background.root.mainloop()

    def searchFood(self, goalCoord):
        while (1):
            self.snake.autoMove(goalCoord)
            #            time.sleep(0.5)
            #            self.snake.move()
            #            print self.snake.snake_coord[0], "  ", self.food.coord
            if self.snake.snake_coord[0] == self.food.coord:
                print "eat"
                self.snake.eat()
                self.food.eaten(self.snake.snake_coord)
                goalCoord = self.food.coord
                time.sleep(0.2)
    def writeStatus(self):
        for x in range(game_height):
            for y in range(game_width):
                self.ws.cell(row=self.data_number, column=game_height * y + x + 1).value = 0
#        print "x", self.snake.snake_coord[0][0]/20, "y", self.snake.snake_coord[0][1]/20
        self.ws.cell(row=self.data_number, column=game_height * self.snake.snake_coord[0][1]/20 + self.snake.snake_coord[0][0]/20 + 1).value = 1
        for i in range(1, self.snake.number):
            self.ws.cell(row=self.data_number, column=game_height * self.snake.snake_coord[i][1]/20 + self.snake.snake_coord[i][0]/20 + 1).value = 2


class Node():
    def __init__(self, x, y):
        self.x = x
        self.y = y
        self.parent = -1
        self.value = -1
        self.visted = 0
        self.id = 0

    def updateValue(self):
        pass


class Pathplan():
    def __init__(self, startCoord, goalCoord, obstacleArray):
        self.root = Tk()
        self.cv = Canvas(self.root, bg='white', height=500, width=600)

        self.start_x = startCoord[0]
        self.start_y = startCoord[1]
        start = self.cv.create_rectangle(self.start_x, self.start_y, self.start_x + 20, self.start_y + 20, fill='green')

        self.goal_x = goalCoord[0]
        self.goal_y = goalCoord[1]
        goal = self.cv.create_rectangle(self.goal_x, self.goal_y, self.goal_x + 20, self.goal_y + 20, fill='yellow')

        #        obstacle_x = 20 * random.randint(2,10)
        #        obstacle_y = 20 * random.randint(2,10)
        self.obstacle_array = obstacleArray
        for i in range(len(obstacleArray)):
            self.cv.create_rectangle(obstacleArray[i][0], obstacleArray[i][1], obstacleArray[i][0] + 20,
                                     obstacleArray[i][1] + 20, fill='black')

        self.nodeList = []
        self.coords = []
        self.cv.pack()

        self.optimalPath = []
        print "goal position: ", self.goal_x, self.goal_y
        self.optimalPath.append([self.goal_x, self.goal_y])

    def update(self, startCoord, goalCoord, obstacleArray):
        self.start_x = startCoord[0]
        self.start_y = startCoord[1]
        self.goal_x = goalCoord[0]
        self.goal_y = goalCoord[1]
        self.obstacle_array = obstacleArray
        self.nodeList = []
        self.coords = []
        self.optimalPath = []
        print "goal position: ", self.goal_x, self.goal_y
        self.optimalPath.append([self.goal_x, self.goal_y])

    def findPath(self):
        node = Node(self.start_x, self.start_y)
        temp = [self.start_x, self.start_y]
        j = 0
        node.visted = 1
        node.value = 0
        node.id = len(self.nodeList)
        self.nodeList.append(node)
        self.coords.append(temp)

        currentID = node.id

        findGoal = 0

        flag = 0
        while findGoal == 0:
            temp_x = self.nodeList[j].x
            temp_y = self.nodeList[j].y
            temp = [temp_x, temp_y]
            currentID = j

            # (temp_x+20, temp_y)
            flag = 0
            node = Node(temp_x + 20, temp_y)
            temp = [node.x, node.y]
            node.parent = currentID
            node.id = currentID + 1
            if node.x == self.goal_x and node.y == self.goal_y:
                findGoal = 1
                node.value = self.nodeList[node.parent].value + 1
                self.coords.append(temp)
                self.nodeList.append(node)
                break
            elif node.x > 0 and node.y > 0:
                if temp not in self.coords:
                    for i in range(len(self.obstacle_array)):
                        if node.x == self.obstacle_array[i][0] and node.y == self.obstacle_array[i][1]:
                            flag = 1
                    # if the node is not an obstacle, then calculate value
                    if flag == 0:
                        node.value = self.nodeList[node.parent].value + 1
                        self.coords.append(temp)
                        self.nodeList.append(node)
                        #                        self.cv.create_rectangle(node.x, node.y, node.x + 20, node.y + 20, fill = 'red')
                else:
                    if node.value > self.nodeList[node.parent].value + 1:
                        node.value = self.nodeList[node.parent].value + 1
            node.visited = 1

            # (temp_x-20, temp_y)
            node = Node(temp_x - 20, temp_y)
            temp = [node.x, node.y]
            node.parent = currentID
            node.id = currentID + 2
            flag = 0
            if node.x == self.goal_x and node.y == self.goal_y:
                findGoal = 1
                node.value = self.nodeList[node.parent].value + 1
                self.coords.append(temp)
                self.nodeList.append(node)
                break
            elif node.x > 0 and node.y > 0:
                if temp not in self.coords:
                    for i in range(len(self.obstacle_array)):
                        if node.x == self.obstacle_array[i][0] and node.y == self.obstacle_array[i][1]:
                            flag = 1
                    # if the node is not an obstacle, then calculate value
                    if flag == 0:
                        node.value = self.nodeList[node.parent].value + 1
                        self.nodeList.append(node)
                        self.coords.append(temp)
                        #                        self.cv.create_rectangle(node.x, node.y, node.x + 20, node.y + 20, fill = 'red')
                else:
                    if node.value > self.nodeList[node.parent].value + 1:
                        node.value = self.nodeList[node.parent].value + 1
            node.visited = 1

            # (temp_x, temp_y + 20)
            node = Node(temp_x, temp_y + 20)
            temp = [node.x, node.y]
            node.parent = currentID
            node.id = currentID + 3
            flag = 0
            if node.x == self.goal_x and node.y == self.goal_y:
                findGoal = 1
                node.value = self.nodeList[node.parent].value + 1
                self.coords.append(temp)
                self.nodeList.append(node)
                break
            elif node.x > 0 and node.y > 0:
                if temp not in self.coords:
                    for i in range(len(self.obstacle_array)):
                        if node.x == self.obstacle_array[i][0] and node.y == self.obstacle_array[i][1]:
                            flag = 1
                    # if the node is not an obstacle, then calculate value
                    if flag == 0:
                        node.value = self.nodeList[node.parent].value + 1
                        self.nodeList.append(node)
                        self.coords.append(temp)
                        #                        self.cv.create_rectangle(node.x, node.y, node.x + 20, node.y + 20, fill = 'red')
                else:
                    if node.value > self.nodeList[node.parent].value + 1:
                        node.value = self.nodeList[node.parent].value + 1
            node.visited = 1

            # (temp_x, temp_y - 20)
            node = Node(temp_x, temp_y - 20)
            temp = [node.x, node.y]
            node.parent = currentID
            node.id = currentID + 4
            flag = 0
            if node.x == self.goal_x and node.y == self.goal_y:
                findGoal = 1
                node.value = self.nodeList[node.parent].value + 1
                self.coords.append(temp)
                self.nodeList.append(node)
                break
            elif node.x > 0 and node.y > 0:
                if temp not in self.coords:
                    for i in range(len(self.obstacle_array)):
                        if node.x == self.obstacle_array[i][0] and node.y == self.obstacle_array[i][1]:
                            flag = 1
                    # if the node is not an obstacle, then calculate value
                    if flag == 0:
                        node.value = self.nodeList[node.parent].value + 1
                        self.nodeList.append(node)
                        self.coords.append(temp)
                        #                        self.cv.create_rectangle(node.x, node.y, node.x + 20, node.y + 20, fill = 'red')
                else:
                    if node.value > self.nodeList[node.parent].value + 1:
                        node.value = self.nodeList[node.parent].value + 1
            node.visited = 1

            j = j + 1
        # self.cv.pack()


        length = len(self.nodeList)
        index = length - 1
        while not self.nodeList[index].parent == 0:
            index = self.nodeList[index].parent
            self.optimalPath.append([self.nodeList[index].x, self.nodeList[index].y])
            self.cv.create_rectangle(self.nodeList[index].x, self.nodeList[index].y, self.nodeList[index].x + 20,
                                     self.nodeList[index].y + 20, fill='red')

            #       for k in range(len(self.optimalPath)):
        #            print "path ", k, "  :",self.optimalPath[k]
        #            time.sleep(0.2)
        #            self.cv.create_rectangle(self.optimalPath[k][0], self.optimalPath[k][1], self.optimalPath[k][0] + 20, self.optimalPath[k][1] + 20, fill = 'green')

        self.cv.pack()
        return self.optimalPath

    def drawPath(self):
        #        self.root.mainloop()
        pass


if __name__ == '__main__':
    print "OK"
    background = Background()
    background.draw()

    game = SnakeGame()